import oracledb
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import getpass
import os

def get_patch_info(connection):
    sql = """
    SELECT *
    FROM (
      SELECT install_id,
             patch_id,
             patch_type,
             action,
             status,
             action_time,
             description,
             CASE
               WHEN LOWER(description) LIKE '%ojvm%' THEN 'OJVM'
               WHEN LOWER(description) LIKE '%ocw%' THEN 'OCW'
               WHEN LOWER(description) LIKE '%database release update%' 
                    OR LOWER(description) LIKE '%windows database bundle patch%' THEN 'DBRU'
               ELSE 'OTHER'
             END AS patch_category,
             ROW_NUMBER() OVER (
               PARTITION BY 
                 CASE
                   WHEN LOWER(description) LIKE '%ojvm%' THEN 'OJVM'
                   WHEN LOWER(description) LIKE '%ocw%' THEN 'OCW'
                   WHEN LOWER(description) LIKE '%database release update%' 
                        OR LOWER(description) LIKE '%windows database bundle patch%' THEN 'DBRU'
                   ELSE 'OTHER'
                 END
               ORDER BY action_time DESC
             ) AS rn
      FROM dba_registry_sqlpatch
      WHERE action = 'APPLY' AND status = 'SUCCESS'
    )
    WHERE rn = 1 AND patch_category IN ('DBRU', 'OJVM', 'OCW')
    """
    cursor = connection.cursor()
    cursor.execute(sql)
    rows = cursor.fetchall()
    columns = [col[0].lower() for col in cursor.description]
    cursor.close()

    patch_info = {"DBRU": None, "OJVM": None, "OCW": None}
    for row in rows:
        row_data = dict(zip(columns, row))
        patch_type = row_data["patch_category"]
        patch_info[patch_type] = {
            "action_time": row_data["action_time"].strftime("%Y-%m-%d %H:%M:%S") if row_data["action_time"] else "",
            "patch_id": row_data["patch_id"],
            "patch_type": row_data["patch_type"],
            "status": row_data["status"],
            "description": row_data["description"]
        }
    return patch_info

def export_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Oracle Patch Info"

    headers = ["Host", "Port", "SID", "Patch Type", "Patch ID", "Status", "Action Time", "Comments"]

    # Header at row 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"  # Freeze above row 3 so rows 1-2 stay fixed

    row_num = 3  # Start writing data from row 3

    for db in data:
        host = db["host"]
        port = db["port"]
        sid = db["sid"]
        patches = db["patches"]

        for patch_type in ["DBRU", "OJVM", "OCW"]:
            patch = patches.get(patch_type)
            values = [
                host,
                port,
                sid,
                patch_type,
                patch.get("patch_id", "") if patch else "",
                patch.get("status", "") if patch else "",
                patch.get("action_time", "") if patch else "",
                patch.get("description", "") if patch else ""
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center")
            row_num += 1

        # Add a blank row (or separator) after each DB entry
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        sep_cell = ws.cell(row=row_num, column=1, value="")
        sep_cell.alignment = Alignment(horizontal="center")
        row_num += 1

    wb.save(filename)
    print(f"\n✅ Excel file generated: {filename}")

def main():
    db_connections = []

    choice = input("Choose input method:\n1. Enter DB connections manually\n2. Load DB connections from .txt file\nEnter 1 or 2: ").strip()

    if choice == "1":
        db_list_input = input("Enter DB connections (host:port:sid, comma-separated): ").strip()
        for item in db_list_input.split(','):
            parts = item.strip().split(':')
            if len(parts) == 3:
                db_connections.append({
                    "host": parts[0],
                    "port": parts[1],
                    "sid": parts[2]
                })
            else:
                print(f"⚠️ Ignoring invalid connection format: {item}")
    elif choice == "2":
        txt_path = input("Enter path to the .txt file: ").strip()
        try:
            with open(txt_path, 'r') as file:
                lines = file.readlines()
            for line in lines:
                parts = line.strip()
                if not line:
                    continue  # skip blank lines silently
                parts = line.split(':')
                if len(parts) == 3:
                    db_connections.append({
                        "host": parts[0],
                        "port": parts[1],
                        "sid": parts[2]
                    })
                else:
                    print(f"⚠️ Invalid line format in file: {line.strip()}")
        except Exception as e:
            print(f"❌ Failed to read file {txt_path}: {e}")
            return  # Exit main due to failure reading file
    else:
        print("⚠️ Invalid choice, exiting.")
        return  # Exit main

    # Now db_connections is ready for processing
    print(f"Found {len(db_connections)} DB connections to process.")
    

    # db_user = input("Enter Oracle DB username: ")
    # db_pass = getpass.getpass("Enter Oracle DB password: ")

    # Hardcoded credentials (you can change to input if you want)
    db_user = "sys as sysdba"
    db_pass = "Ebslabs123"

    collected_data = []
    for db in db_connections:
        dsn = oracledb.makedsn(db["host"], db["port"], sid=db["sid"])
        print(f"\n🔗 Connecting to {db['host']}:{db['port']}/{db['sid']} ...")

        is_sysdba = ' as sysdba' in db_user.lower()
        clean_user = db_user.lower().replace(' as sysdba', '').strip()

        try:
            conn = oracledb.connect(
                user=clean_user,
                password=db_pass,
                dsn=dsn,
                mode=oracledb.AUTH_MODE_SYSDBA if is_sysdba else None
            )
            patches = get_patch_info(conn)
            collected_data.append({
                "host": db["host"],
                "port": db["port"],
                "sid": db["sid"],
                "patches": patches
            })
            conn.close()
        except oracledb.DatabaseError as e:
            print(f"❌ Failed to connect or query {db['host']}:{db['port']}/{db['sid']}: {e}")

    if collected_data:
        filename = f"oracle_db_patch_info_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_to_excel(collected_data, filename)
    else:
        print("⚠️ No data collected, no Excel file generated.")

if __name__ == "__main__":
    main()
